"""
tools/generate_xlsx.py — Gera arquivo .xlsx a partir dos dados do relatório.

Usa openpyxl. Inclui cabeçalho com nome da escola, título do relatório,
tabela com dados e rodapé com data de geração.

Retorna o file_id para download via GET /report/download/{file_id}.
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any

import structlog
from langchain_core.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.file_storage import save_file

logger = structlog.get_logger(__name__)

_HEADER_FILL   = PatternFill("solid", fgColor="4338CA")   # roxo Eleve
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT    = Font(bold=True, size=13)
_ALT_FILL      = PatternFill("solid", fgColor="F5F3FF")   # roxo claro alternado


@tool
async def generate_xlsx(
    data: list,
    columns: list,
    title: str,
    school_id: str,
    school_name: str = "Escola",
    sa_token: str = "",
    **kwargs,
) -> str:
    """
    Gera relatório .xlsx e salva em file storage temporário.

    Args:
        data:        Lista de dicionários com os dados.
        columns:     Lista de colunas a exibir (chaves dos dicts).
        title:       Título do relatório.
        school_id:   ID da escola (usado no nome do arquivo).
        school_name: Nome da escola para o cabeçalho.

    Returns:
        JSON string com { file_id: str, rows: int }
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = title[:30]

        # ── Cabeçalho da escola ───────────────────────────────────────
        ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
        ws["A1"] = school_name
        ws["A1"].font      = _TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells(f"A2:{get_column_letter(len(columns))}2")
        ws["A2"] = title
        ws["A2"].font      = Font(bold=True, size=11)
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells(f"A3:{get_column_letter(len(columns))}3")
        ws["A3"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} pelo Nicodemus ADM"
        ws["A3"].font      = Font(italic=True, size=9, color="888888")
        ws["A3"].alignment = Alignment(horizontal="center")

        ws.append([])  # linha vazia

        # ── Cabeçalhos da tabela ──────────────────────────────────────
        header_row = 5
        for col_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=header_row, column=col_idx, value=_format_header(col_name))
            cell.fill      = _HEADER_FILL
            cell.font      = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Dados ─────────────────────────────────────────────────────
        for row_idx, row in enumerate(data, start=header_row + 1):
            alt = (row_idx % 2 == 0)
            for col_idx, col_name in enumerate(columns, start=1):
                value = _get_value(row, col_name)
                cell  = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="center")
                if alt:
                    cell.fill = _ALT_FILL

        # ── Ajuste de largura das colunas ─────────────────────────────
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = max(
                len(_format_header(col_name)),
                max((len(str(_get_value(row, col_name) or "")) for row in data), default=0),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # ── Serializa e salva ─────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        file_id = save_file(school_id, buf.getvalue(), "xlsx")

        logger.info("generate_xlsx.ok", file_id=file_id, rows=len(data))
        return json.dumps({"file_id": file_id, "rows": len(data)})

    except Exception as exc:
        logger.error("generate_xlsx.error", error=str(exc))
        return json.dumps({"error": str(exc)})


def _format_header(col_name: str) -> str:
    return col_name.replace("_", " ").title()


def _get_value(row: dict, col_name: str) -> Any:
    val = row.get(col_name)
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    return val
