"""
routers/upload.py — Endpoint de upload contextual de arquivo do Nicodemus ADM.

Fluxo (15 passos):
  1.  Valida MIME do arquivo (415 se tipo não suportado)
  2.  Verifica tamanho ≤ 10 MB (413 se exceder)
  3.  Detecta categoria do arquivo pelo nome
      (calendario_escolar, matricula, financeiro, turma, professores, notas, geral)
  4.  Extrai conteúdo do arquivo:
        PDF/imagem → GPT-4o Vision (base64)
        XLSX       → openpyxl
        CSV        → módulo csv
        DOCX       → python-docx
        TXT        → decode utf-8
  5.  Constrói user_text: "[ARQUIVO ENVIADO: {nome} | categoria: {cat}]\n{conteúdo}\n\n{msg}"
  6.  Carrega/retoma sessão via SessionService
  7.  Carrega contexto Redis
  8.  Persiste mensagem do usuário no banco
  9.  Monta NicoState
  10. Invoca nico_graph.ainvoke(state)
  11. Extrai reply do estado final
  12. Persiste mensagem do assistente no banco
  13. Atualiza contexto Redis com o par user/assistant
  14. Incrementa report_count se houve geração de arquivo
  15. Retorna ChatResponse
"""
from __future__ import annotations

import base64
import csv
import io
import json
import re

import structlog
from docx import Document
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openai import AsyncOpenAI
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from agent.nico_agent import nico_graph
from agent.state import NicoState
from core import memory
from core.auth import CurrentUser, get_current_user
from core.database import get_session
from core.settings import settings
from schemas.session_types import ChatResponse
from services.session_service import SessionService

logger = structlog.get_logger(__name__)

router = APIRouter()

# ── Configuração ───────────────────────────────────────────────────────────────

ALLOWED_MIMES: set[str] = {
    "application/pdf",
    "image/jpeg",
    "image/jpg",
    "image/png",
    "image/webp",
    "image/gif",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/csv",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "text/plain",
}

VISION_MIMES: set[str] = {
    "application/pdf",
    "image/jpeg",
    "image/jpg",
    "image/png",
    "image/webp",
    "image/gif",
}

MAX_BYTES = 10 * 1024 * 1024  # 10 MB

_CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "calendario_escolar": ["calendario", "calendário", "calendar"],
    "matricula": ["matricula", "matrícula", "enrollment"],
    "financeiro": ["financeiro", "financeira", "finance", "boleto", "pagamento", "payment"],
    "turma": ["turma", "classe", "class", "alunos", "students"],
    "professores": ["professor", "professores", "teacher", "docente", "docentes"],
    "notas": ["notas", "nota", "grades", "grade", "boletim"],
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def _detect_category(filename: str) -> str:
    """Retorna a categoria do arquivo com base no nome. Padrão: 'geral'."""
    name_lower = filename.lower()
    name_clean = re.sub(r"[^a-záéíóúàâêôãõça-z0-9 _-]", "", name_lower)
    for category, keywords in _CATEGORY_KEYWORDS.items():
        if any(kw in name_clean for kw in keywords):
            return category
    return "geral"


async def _extract_via_vision(content: bytes, mime_type: str) -> str:
    """Extrai texto de PDF ou imagem usando GPT-4o Vision."""
    client = AsyncOpenAI(api_key=settings.openai_api_key)
    b64 = base64.b64encode(content).decode()

    try:
        response = await client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{b64}",
                                "detail": "high",
                            },
                        },
                        {
                            "type": "text",
                            "text": (
                                "Extraia todo o texto e dados relevantes deste documento de "
                                "gestão escolar. Organize as informações de forma estruturada "
                                "e completa. Inclua todos os números, datas, nomes e valores."
                            ),
                        },
                    ],
                }
            ],
            max_tokens=4096,
        )
        return response.choices[0].message.content or ""
    except Exception as exc:
        logger.warning("upload.vision_error", error=str(exc), mime=mime_type)
        raise HTTPException(
            status_code=422,
            detail=f"Não foi possível extrair o conteúdo do arquivo: {exc}",
        )


def _extract_xlsx(content: bytes) -> str:
    """Extrai dados de planilha XLSX usando openpyxl."""
    try:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        lines: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== Aba: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                row_values = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in row_values):
                    lines.append("\t".join(row_values))
        return "\n".join(lines)
    except Exception as exc:
        logger.warning("upload.xlsx_error", error=str(exc))
        raise HTTPException(status_code=422, detail=f"Erro ao processar planilha XLSX: {exc}")


def _extract_csv(content: bytes) -> str:
    """Extrai dados de arquivo CSV."""
    try:
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return "\n".join("\t".join(row) for row in reader)
    except Exception as exc:
        logger.warning("upload.csv_error", error=str(exc))
        raise HTTPException(status_code=422, detail=f"Erro ao processar arquivo CSV: {exc}")


def _extract_docx(content: bytes) -> str:
    """Extrai texto de arquivo DOCX usando python-docx."""
    try:
        doc = Document(io.BytesIO(content))
        return "\n".join(para.text for para in doc.paragraphs if para.text.strip())
    except Exception as exc:
        logger.warning("upload.docx_error", error=str(exc))
        raise HTTPException(status_code=422, detail=f"Erro ao processar arquivo DOCX: {exc}")


async def _extract_file_content(filename: str, content: bytes, mime_type: str) -> str:  # noqa: ARG001
    """Extrai o conteúdo textual do arquivo conforme o tipo MIME."""
    if mime_type in VISION_MIMES:
        return await _extract_via_vision(content, mime_type)
    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return _extract_xlsx(content)
    if mime_type == "text/csv":
        return _extract_csv(content)
    if mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return _extract_docx(content)
    # TXT e outros texto puro
    return content.decode("utf-8", errors="replace")


def _extract_file_id(messages: list[dict]) -> str | None:
    """Extrai o file_id do primeiro resultado de tool que contenha esse campo."""
    for msg in messages:
        if msg.get("role") == "tool":
            try:
                data = json.loads(msg.get("content", "{}"))
                if isinstance(data, dict) and data.get("file_id"):
                    return str(data["file_id"])
            except (json.JSONDecodeError, AttributeError):
                pass
    return None


# ── Endpoint ──────────────────────────────────────────────────────────────────

@router.post("/upload", response_model=ChatResponse)
async def upload_file(
    file: UploadFile = File(...),
    session_id: str = Form(...),
    message: str = Form(default=""),
    user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_session),
):
    """
    Recebe um arquivo do chat do Nicodemus ADM, extrai seu conteúdo
    e processa como mensagem contextual para o agente.

    Suporta: PDF, imagens (JPEG, PNG, WEBP, GIF), XLSX, CSV, DOCX, TXT.
    Limite: 10 MB por arquivo.
    Requer sessão ativa — crie via POST /sessions/.
    """
    # ── Passo 1: Validação de MIME ────────────────────────────────────────────────
    mime_type = file.content_type or "application/octet-stream"
    if mime_type not in ALLOWED_MIMES:
        raise HTTPException(
            status_code=415,
            detail=(
                f"Tipo de arquivo não suportado: {mime_type}. "
                "Aceitos: PDF, imagens (JPEG/PNG/WEBP/GIF), XLSX, CSV, DOCX, TXT."
            ),
        )

    # ── Passo 2: Verificação de tamanho ──────────────────────────────────────────
    content = await file.read()
    if len(content) > MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail=(
                f"Arquivo excede o limite de 10 MB "
                f"({len(content) / (1024 * 1024):.1f} MB recebidos)."
            ),
        )

    filename = file.filename or "arquivo_sem_nome"

    # ── Passo 3: Detecção de categoria ─────────────────────────────────────────
    category = _detect_category(filename)
    logger.info(
        "upload.received",
        filename=filename,
        mime=mime_type,
        category=category,
        size_bytes=len(content),
    )

    # ── Passo 4: Extração de conteúdo ──────────────────────────────────────────
    extracted_text = await _extract_file_content(filename, content, mime_type)

    # ── Passo 5: Construção do user_text ───────────────────────────────────────
    header = f"[ARQUIVO ENVIADO: {filename} | categoria: {category}]"
    user_text_parts = [header, extracted_text]
    if message.strip():
        user_text_parts.append(f"\n{message.strip()}")
    user_text = "\n".join(user_text_parts)

    # ── Passo 6: Carrega sessão ─────────────────────────────────────────────────
    try:
        session = await SessionService.get_or_resume(db, session_id, user.user_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))

    if session.status == "completed":
        raise HTTPException(
            status_code=400,
            detail="Sessão encerrada. Crie uma nova sessão para continuar.",
        )

    # ── Passo 7: Contexto Redis ────────────────────────────────────────────────
    context = await memory.get_context(session_id)
    user_msg = {"role": "user", "content": user_text}
    messages_for_agent = context + [user_msg]

    # ── Passo 8: Persiste mensagem do usuário ──────────────────────────────────────
    await SessionService.add_message(db, session, "user", user_text)

    logger.info(
        "upload.invoke",
        session_id=session_id,
        user_id=user.user_id,
        filename=filename,
        category=category,
        msg_count=len(messages_for_agent),
    )

    # ── Passos 9–10: Monta NicoState e invoca agente ──────────────────────────────
    initial_state: NicoState = {
        "user_id":      user.user_id,
        "school_id":    user.school_id,
        "sa_token":     user.sa_token,
        "role":         user.role,
        "user_name":    user.name,
        "session_id":   session_id,
        "messages":     messages_for_agent,
        "user_message": user_text,
        "tool_calls":   [],
        "response":     "",
        "error":        None,
    }

    final_state: NicoState = await nico_graph.ainvoke(initial_state)

    # ── Passo 11: Extrai reply ────────────────────────────────────────────────────
    reply = final_state.get("response") or ""
    if not reply:
        reply = "Não consegui processar o arquivo. Tente novamente."

    all_messages = final_state.get("messages", [])
    file_id = _extract_file_id(all_messages)
    file_url = f"/report/download/{file_id}" if file_id else None

    # ── Passo 12: Persiste mensagem do assistente ─────────────────────────────────
    await SessionService.add_message(
        db,
        session,
        "assistant",
        reply,
        metadata={"file_id": file_id} if file_id else None,
    )

    # ── Passo 13: Atualiza Redis ─────────────────────────────────────────────────
    assistant_msg = {"role": "assistant", "content": reply}
    await memory.append_turn(session_id, user_msg, assistant_msg)

    # ── Passo 14: Incrementa report_count se houve geração de arquivo ──────────────
    if file_id:
        await SessionService.increment_report_count(db, session)

    logger.info(
        "upload.ok",
        session_id=session_id,
        user_id=user.user_id,
        filename=filename,
        has_file=file_id is not None,
    )

    # ── Passo 15: Retorna resposta ────────────────────────────────────────────────
    return ChatResponse(
        session_id=session_id,
        reply=reply,
        file_id=file_id,
        file_url=file_url,
    )
